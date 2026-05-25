# -*- coding: utf-8 -*-
"""
专利数据清洗融合 v5.0 (Electron版)
支持命令行参数: --data-dir <path>
从指定数据目录读取三个爬虫的输出文件，合并清洗后输出 cleaned_data.json
"""

import os, re, json, csv, sys, argparse
from datetime import datetime, date
from collections import Counter
import glob


def safe_str(v):
    return str(v or '')


# ========== 1. 读取Excel ==========
def read_new_excel(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        headers[c] = str(ws.cell(1, c).value or '').strip()

    def find_col(keywords):
        for c, h in headers.items():
            for kw in keywords:
                if h == kw or kw in h:
                    return c - 1
        return None

    col_map = {
        'applyId': find_col(['申请号']),
        'pubId': find_col(['公开（公告）号', '公开公告号', '公开号']),
        'applyDate': find_col(['申请日']),
        'pubDate': find_col(['公开（公告）日', '公开公告日', '公开日']),
        'classification': find_col(['IPC分类号', 'IPC']),
        'applicant': find_col(['申请（专利权）人', '申请人', '专利权人']),
        'inventor': find_col(['发明人']),
        'title': find_col(['发明名称', '专利名称', '名称']),
        'priority': find_col(['优先权号']),
        'patentAgent': find_col(['专利代理师', '代理人']),
        'patentAgency': find_col(['专利代理机构', '代理机构']),
        'patentType': find_col(['文献类型']),
        'abstract': find_col(['摘要']),
        'zipcode': find_col(['邮编', '邮政编码']),
    }

    addr_col = find_col(['地址'])
    if addr_col is None:
        for c, h in headers.items():
            if '地址' in h:
                addr_col = c - 1
                break

    records = []
    for r in range(2, ws.max_row + 1):
        row = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        rec = {}
        for k, ci in col_map.items():
            if ci is not None and ci < len(row):
                rec[k] = row[ci]
        if addr_col is not None and addr_col < len(row):
            rec['address'] = row[addr_col]
        records.append(rec)

    print(f"  Excel共读取 {len(records)} 条记录")
    wb.close()
    return records


# ========== 2. 读取CSV（专利公告）==========
def read_patent_csv(path):
    records = []
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append(dict(row))
    print(f"  CSV共读取 {len(records)} 条记录: {path}")
    return records


# ========== 3. 读取天眼查JSON ==========
def read_tianyan_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, list):
        print(f"  JSON共读取 {len(data)} 条记录 (列表)")
        return data
    elif isinstance(data, dict) and 'data' in data:
        items = data['data']
        print(f"  JSON共读取 {len(items)} 条记录 (data字段)")
        return items
    else:
        print(f"  JSON读取: 字典类型，键: {list(data.keys())[:10]}")
        return [data]


# ========== 4. 标准化 ==========
def normalize_date(d):
    if not d or d in ('无', '未知', 'None', ''):
        return None
    d = str(d).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日',
                '%Y%m%d', '%Y-%m-%d %H:%M:%S', '%Y.%m'):
        try:
            return datetime.strptime(d, fmt).strftime('%Y-%m-%d')
        except:
            pass
    m = re.match(r'(\d{4})[-/.]?(\d{1,2})[-/.]?(\d{1,2})?', d)
    if m:
        y, mo, da = m.group(1), m.group(2), m.group(3)
        if da:
            return f"{y}-{mo.zfill(2)}-{da.zfill(2)}"
        return f"{y}-{mo.zfill(2)}"
    return d


def normalize_patent_type(t):
    if not t:
        return '未知'
    t = str(t).strip()
    mapping = {
        '发明': '发明专利', '发明公布': '发明专利', '发明授权': '发明专利',
        '发明专利申请': '发明专利', '1': '发明专利', 'A': '发明专利',
        '实用新型': '实用新型', '实用新型专利': '实用新型', '2': '实用新型', 'U': '实用新型',
        '外观设计': '外观设计', '3': '外观设计', 'D': '外观设计',
        '发明授权': '发明授权', 'B': '发明授权',
    }
    for k, v in mapping.items():
        if k in t:
            return v
    return t


def parse_ipc(code):
    if not code:
        return []
    parts = re.split(r'[;；,，、\s]+', str(code))
    return [p.strip() for p in parts if p.strip()]


# ========== 5. 合并去重 ==========
def merge_records(records, id_field='applyId'):
    seen = set()
    unique = []
    for r in records:
        key = r.get(id_field, '')
        if key and key not in seen:
            seen.add(key)
            unique.append(r)
    print(f"  去重: {len(records)} -> {len(unique)}")
    return unique


# ========== 6. 主函数 ==========
def main():
    parser = argparse.ArgumentParser(description='专利数据清洗融合')
    parser.add_argument('--data-dir', required=True, help='用户数据目录（包含爬虫输出文件）')
    args = parser.parse_args()

    data_dir = args.data_dir
    print(f"=" * 60)
    print(f"专利数据清洗融合 v5.0")
    print(f"数据目录: {data_dir}")
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"=" * 60)

    all_records = []

    # ----- 6.1 读取爬虫A（CSV）-----
    csv_files = glob.glob(os.path.join(data_dir, "专利数据_v2_*.csv"))
    csv_files += glob.glob(os.path.join(data_dir, "*.csv"))
    if csv_files:
        print(f"\n[爬虫A] 发现CSV文件: {len(csv_files)}个")
        for fpath in csv_files:
            print(f"  读取: {os.path.basename(fpath)}")
            try:
                recs = read_patent_csv(fpath)
                for r in recs:
                    r['_source'] = 'patent_announce'
                all_records.extend(recs)
            except Exception as e:
                print(f"  [错误] 读取CSV失败: {e}")
    else:
        print(f"\n[爬虫A] 未找到CSV文件")

    # ----- 6.2 读取爬虫B（Excel/xlsx）-----
    excel_files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if excel_files:
        print(f"\n[爬虫B] 发现Excel文件: {len(excel_files)}个")
        for fpath in excel_files:
            print(f"  读取: {os.path.basename(fpath)}")
            try:
                recs = read_new_excel(fpath)
                for r in recs:
                    r['_source'] = 'cnipa'
                all_records.extend(recs)
            except Exception as e:
                print(f"  [错误] 读取Excel失败: {e}")
    else:
        print(f"\n[爬虫B] 未找到Excel文件")

    # ----- 6.3 读取爬虫C（JSON或xlsx）-----
    json_files = glob.glob(os.path.join(data_dir, "*天眼查*.json"))
    json_files += glob.glob(os.path.join(data_dir, "01_专利数据.json"))
    if json_files:
        print(f"\n[爬虫C] 发现JSON文件: {len(json_files)}个")
        for fpath in json_files:
            print(f"  读取: {os.path.basename(fpath)}")
            try:
                recs = read_tianyan_json(fpath)
                for r in recs:
                    r['_source'] = 'tianyancha'
                all_records.extend(recs)
            except Exception as e:
                print(f"  [错误] 读取JSON失败: {e}")
    else:
        # 也可能天眼查输出的是xlsx
        tianyan_xlsx = [f for f in excel_files if '天眼查' in f]
        if tianyan_xlsx:
            print(f"\n[爬虫C] 发现天眼查Excel文件: {len(tianyan_xlsx)}个")
            for fpath in tianyan_xlsx:
                try:
                    recs = read_new_excel(fpath)
                    for r in recs:
                        r['_source'] = 'tianyancha'
                    all_records.extend(recs)
                except Exception as e:
                    print(f"  [错误] 读取失败: {e}")
        else:
            print(f"\n[爬虫C] 未找到天眼查文件")

    print(f"\n{'=' * 60}")
    print(f"原始数据总量: {len(all_records)}")

    # ----- 6.4 标准化 -----
    print(f"\n[标准化处理]...")
    normalized = []
    for r in all_records:
        nr = {}
        for k, v in r.items():
            nr[k] = safe_str(v)

        # 标准化日期
        for dk in ['applyDate', 'pubDate', '申请日', '公开日', '公开公告日', '申请日期']:
            if dk in nr:
                nr[dk] = normalize_date(nr[dk])

        # 标准化类型
        for tk in ['patentType', '专利类型', '文献类型', '类型']:
            if tk in nr:
                nr[tk] = normalize_patent_type(nr[tk])

        # 标准化IPC
        for ik in ['classification', 'IPC分类号', 'IPC', '分类号']:
            if ik in nr and nr[ik]:
                nr[ik + '_list'] = parse_ipc(nr[ik])

        normalized.append(nr)

    # ----- 6.5 去重 -----
    print(f"\n[去重处理]...")
    deduped = merge_records(normalized, id_field='申请号')
    if len(deduped) < 10:
        deduped = merge_records(normalized, id_field='applyId')
    if len(deduped) < 10:
        deduped = merge_records(normalized, id_field='pubId')
    if len(deduped) < 10:
        deduped = merge_records(normalized, id_field='公开（公告）号')

    print(f"最终数据: {len(deduped)} 条")

    # ----- 6.6 输出JSON -----
    output_path = os.path.join(data_dir, "cleaned_data.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(deduped, f, ensure_ascii=False, indent=2)
    print(f"\n已输出: {output_path}")

    # 返回统计信息（stdout输出JSON供Electron读取）
    result = {
        "success": True,
        "total_raw": len(all_records),
        "total_cleaned": len(deduped),
        "output_file": output_path,
        "timestamp": datetime.now().isoformat(),
        "sources": {
            "patent_announce": len([r for r in all_records if r.get('_source') == 'patent_announce']),
            "cnipa": len([r for r in all_records if r.get('_source') == 'cnipa']),
            "tianyancha": len([r for r in all_records if r.get('_source') == 'tianyancha']),
        }
    }
    print(f"\n[RESULT]{json.dumps(result, ensure_ascii=False)}[/RESULT]")

    return deduped


if __name__ == "__main__":
    main()