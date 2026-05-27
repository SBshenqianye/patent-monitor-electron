# -*- coding: utf-8 -*-
"""
专利数据清洗融合脚本（直接读取原始文件）
输入：../temp_data/中国专利公布公告网/*.csv
      ../temp_data/天眼查/*.xlsx
      ../temp_data/专利检索及分析网/*.xlsx
输出：./cleaned_data/专利数据_清洗融合.json + .xlsx
日志：./log/数据清洗/
"""

import os
import argparse
import re
import json
import csv
import sys
import calendar
import logging
from datetime import datetime, date
from glob import glob
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== 日志配置 ====================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


def setup_file_logging(data_dir):
    """添加文件日志"""
    log_dir = os.path.join(data_dir, "log", "数据清洗")
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"clean_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.INFO)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logging.getLogger().addHandler(fh)
    logger.info(f"日志文件: {log_file}")


sys.stdout.reconfigure(encoding='utf-8')

# ==================== 工具函数 ====================
def normalize_apply_id(raw_id):
    if not raw_id:
        return ""
    raw = str(raw_id).strip()
    v1 = re.sub(r'^CN', '', raw).strip()
    return v1.replace('.', '')

def parse_apply_date(date_str):
    if not date_str:
        return ""
    s = str(date_str).strip()[:10]
    for fmt in ["%Y-%m-%d", "%Y.%m.%d", "%Y年%m月%d日", "%Y/%m/%d"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s

def infer_patent_type(title):
    """从标题推断专利类型"""
    if not title:
        return ''
    if '发明' in title:
        return '发明'
    if '实用新型' in title or '实用' in title:
        return '实用新型'
    if '外观设计' in title or '外观' in title:
        return '外观设计'
    return ''

def compute_expiry(apply_date_str, patent_type, title=''):
    """
    计算专利到期日，对外观设计考虑新旧法
    """
    if not apply_date_str:
        return "", -1
    d = parse_apply_date(apply_date_str)
    if not d:
        return "", -1
    try:
        ad = datetime.strptime(d, "%Y-%m-%d").date()
    except:
        return "", -1

    # 确定类型（优先使用传入的类型，否则从标题推断）
    ptype = patent_type if patent_type else infer_patent_type(title)
    if not ptype:
        # 默认按发明处理（最长时间）
        years = 20
    elif '发明' in ptype:
        years = 20
    elif '实用' in ptype:
        years = 10
    elif '外观' in ptype:
        if ad >= date(2021, 6, 1):
            years = 15
        else:
            years = 10
    else:
        years = 20

    try:
        expiry = date(ad.year + years, ad.month, ad.day)
    except ValueError:
        last_day = calendar.monthrange(ad.year + years, ad.month)[1]
        expiry = date(ad.year + years, ad.month, min(ad.day, last_day))

    days = (expiry - date.today()).days
    return expiry.strftime("%Y-%m-%d"), days

def clean_title(title):
    if not title:
        return ""
    return re.sub(r'\s+', ' ', str(title)).strip()

def clean_inventor(raw):
    if not raw:
        return ""
    text = str(raw).replace('\n', ' ').replace('\r', '')
    text = re.sub(r'\s*全部\s*', '', text)
    text = re.sub(r'[;；、,，]+', ';', text)
    text = re.sub(r'\s*;\s*', '; ', text).strip('; ')
    return text

def parse_address_field(raw_addr):
    """解析CSV‘地址’大字段，提取地址/分类号/代理/摘要"""
    if not raw_addr:
        return "", "", "", "", ""
    text = str(raw_addr).strip()

    # 截断“事务数据”之后的内容
    idx = text.find("事务数据")
    if idx != -1:
        text = text[:idx]

    # 提取摘要（修复核心）
    abstract = ""
    m = re.search(r'摘要[：:]\s*(.+)', text, re.DOTALL)
    if m:
        abstract = m.group(1)
        # 删除末尾的“全部”及其后的“发明专利申请”等
        abstract = re.sub(r'\s*全部\s*$', '', abstract)
        abstract = re.sub(r'\s*发明专利申请\s*$', '', abstract)
        abstract = re.sub(r'\s+', ' ', abstract).strip()
        text = text[:m.start()]

    # 提取专利代理师
    patent_agent = ""
    m = re.search(r'专利代理师[：:]\s*(.+?)(?=\s*(?:专利代理机构|摘要|$))', text, re.DOTALL)
    if m:
        patent_agent = re.sub(r'\s+', '', m.group(1)).strip()
        text = text[:m.start()]

    # 提取专利代理机构
    patent_agency = ""
    m = re.search(r'专利代理机构[：:]\s*(.+?)(?=\s*(?:专利代理师|摘要|$))', text, re.DOTALL)
    if m:
        patent_agency = re.sub(r'\s+', ' ', m.group(1)).strip()
        text = text[:m.start()]

    # 提取分类号
    classification = ""
    m = re.search(r'分类号[：:]\s*(.+?)(?=\s*(?:专利代理机构|专利代理师|摘要|$))', text, re.DOTALL)
    if m:
        classification = re.sub(r'\s+', ' ', m.group(1)).replace('全部', '').strip('; ')
        text = text[:m.start()]

    # 剩余为地址
    address = re.sub(r'\s*发明专利申请\s*', '', text).strip()
    address = re.sub(r'\s+', '', address)
    return address, classification, patent_agency, patent_agent, abstract


# ==================== 读取各网站原始文件 ====================
def read_publish_csv(folder):
    files = glob(os.path.join(folder, '*.csv'))
    if not files:
        logger.warning("未找到 CSV 文件")
        return []
    latest = max(files, key=os.path.getmtime)
    logger.info(f"读取: {os.path.basename(latest)}")
    records = []
    with open(latest, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_id = row.get('申请号', '')
            aid = normalize_apply_id(raw_id)
            if not aid:
                continue
            addr_raw = row.get('地址', '')
            addr, cls, agency, agent, abstract = parse_address_field(addr_raw)
            raw_title = row.get('标题', '')
            title = re.sub(r'^\[.*?\]\s*', '', raw_title)
            records.append({
                'applyId': aid,
                'title': clean_title(title),
                'applicant': re.sub(r'\s+', ' ', row.get('申请人', '').strip()),
                'inventor': clean_inventor(row.get('发明人', '')),
                'address': addr,
                'classification': cls,
                'patentAgency': agency,
                'patentAgent': agent,
                'abstract': abstract,
                'applyDate': parse_apply_date(row.get('申请日', '')),
                'pubDate': parse_apply_date(row.get('公开（公告）日', '')),
                'source': '中国专利公布公告',
            })
    return records

def read_tianyan_xlsx(folder):
    files = glob(os.path.join(folder, '*.xlsx'))
    if not files:
        logger.warning("未找到 XLSX 文件")
        return []
    latest = max(files, key=os.path.getmtime)
    logger.info(f"读取: {os.path.basename(latest)}")
    wb = openpyxl.load_workbook(latest, data_only=True)
    ws = wb['专利信息']
    records = []
    company_name = ""
    for row in ws.iter_rows(min_row=6, max_row=6, max_col=1, values_only=True):
        if row[0]:
            m = re.search(r'【(.+?)】', str(row[0]))
            if m:
                company_name = m.group(1)

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        raw_id = str(row[4]).strip() if row[4] else ''
        aid = normalize_apply_id(raw_id)
        if not aid:
            continue
        ptype = str(row[2]).strip() if row[2] else ''
        mapped_type = '发明' if '发明' in ptype else '实用新型' if '实用' in ptype else '外观设计' if '外观' in ptype else ptype
        records.append({
            'applyId': aid,
            'title': str(row[1]).strip() if row[1] else '',
            'company': company_name,
            'patentType': mapped_type,
            'legalStatus': str(row[3]).strip() if row[3] else '',
            'applyDate': parse_apply_date(row[5]),
            'pubId': str(row[6]).strip() if row[6] else '',
            'pubDate': parse_apply_date(row[7]),
            'inventor': clean_inventor(row[8]),
            'source': '天眼查',
        })
    wb.close()
    return records

def read_search_xlsx(folder):
    files = glob(os.path.join(folder, '*.xlsx'))
    if not files:
        logger.warning("未找到 XLSX 文件")
        return []
    latest = max(files, key=os.path.getmtime)
    logger.info(f"读取: {os.path.basename(latest)}")
    wb = openpyxl.load_workbook(latest, data_only=True)
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        headers[c] = str(ws.cell(1, c).value or '').strip()
    def find_col(keywords):
        for c, h in headers.items():
            for kw in keywords:
                if h == kw or kw in h:
                    return c - 1
        return None
    col_map = {
        'applyId': find_col(['申请号']),
        'title': find_col(['发明名称', '专利名称', '名称']),
        'pubId': find_col(['公开（公告）号', '公开公告号']),
        'applyDate': find_col(['申请日']),
        'pubDate': find_col(['公开（公告）日', '公开公告日']),
        'classification': find_col(['IPC分类号', 'IPC']),
        'applicant': find_col(['申请（专利权）人', '申请人', '专利权人']),
        'inventor': find_col(['发明人']),
        'patentAgency': find_col(['专利代理机构']),
        'patentAgent': find_col(['专利代理师', '代理人']),
        'abstract': find_col(['摘要']),
        'address': find_col(['地址']),
        'patentType': find_col(['文献类型', '专利类型']),
    }
    records = []
    for r in range(2, ws.max_row + 1):
        raw_id = str(ws.cell(r, (col_map['applyId'] or 0) + 1).value or '').strip()
        aid = normalize_apply_id(raw_id)
        if not aid:
            continue
        def get_cell(key):
            idx = col_map.get(key)
            if idx is None:
                return ''
            return str(ws.cell(r, idx + 1).value or '').strip()
        records.append({
            'applyId': aid,
            'title': get_cell('title').replace('\n', ' '),
            'pubId': get_cell('pubId'),
            'applyDate': parse_apply_date(get_cell('applyDate')),
            'pubDate': parse_apply_date(get_cell('pubDate')),
            'classification': get_cell('classification').replace('全部', '').strip('; '),
            'applicant': get_cell('applicant').replace('\n', ' '),
            'inventor': clean_inventor(get_cell('inventor')),
            'patentAgency': get_cell('patentAgency'),
            'patentAgent': get_cell('patentAgent'),
            'abstract': get_cell('abstract'),
            'address': get_cell('address'),
            'patentType': get_cell('patentType'),
            'source': '专利检索分析网',
        })
    wb.close()
    return records


# ==================== 融合逻辑 ====================
def merge_all(search_data, tianyan_data, publish_data):
    merged = {}
    for rec in search_data:
        merged[rec['applyId']] = dict(rec)
        merged[rec['applyId']]['source'] = '专利检索分析网'

    for rec in tianyan_data:
        aid = rec['applyId']
        if aid in merged:
            item = merged[aid]
            for key in ['company', 'legalStatus', 'patentType']:
                if rec.get(key) and not item.get(key):
                    item[key] = rec[key]
            if rec.get('pubId') and not item.get('pubId'):
                item['pubId'] = rec['pubId']
            if not item.get('source') or '天眼查' not in item['source']:
                item['source'] += '; 天眼查'
        else:
            merged[aid] = dict(rec)
            merged[aid]['source'] = '天眼查'

    for rec in publish_data:
        aid = rec['applyId']
        if aid in merged:
            item = merged[aid]
            for key in ['classification', 'patentAgency', 'patentAgent', 'abstract', 'address', 'applicant']:
                if rec.get(key) and not item.get(key):
                    item[key] = rec[key]
            if rec.get('pubDate') and not item.get('pubDate'):
                item['pubDate'] = rec['pubDate']
            if rec.get('inventor'):
                exist = set(item.get('inventor', '').split('; ')) if item.get('inventor') else set()
                new = set(rec['inventor'].split('; '))
                combined = exist | new
                combined.discard('')
                if combined:
                    item['inventor'] = '; '.join(sorted(combined))
            if '公布公告' not in item.get('source', ''):
                item['source'] += '; 中国专利公布公告'
        else:
            merged[aid] = dict(rec)
            merged[aid]['source'] = '中国专利公布公告'
    return list(merged.values())


# ==================== 最终清洗 ====================
def final_clean(records):
    cleaned = []
    for rec in records:
        item = {
            'applyId': rec['applyId'],
            'title': re.sub(r'\s+', ' ', rec.get('title', '')).strip(),
            'applyDate': rec.get('applyDate', '')[:10] if rec.get('applyDate') else '',
            'pubDate': rec.get('pubDate', '')[:10] if rec.get('pubDate') else '',
            'applicant': re.sub(r'\s+', ' ', rec.get('applicant', '')).strip(),
            'company': rec.get('company', ''),
            'inventor': clean_inventor(rec.get('inventor', '')),
            'address': re.sub(r'\s+', '', rec.get('address', '')).replace('全部', ''),
            'classification': rec.get('classification', '').replace('全部', '').replace(' ', ''),
            'patentAgency': rec.get('patentAgency', ''),
            'patentAgent': rec.get('patentAgent', ''),
            'abstract': rec.get('abstract', ''),
            'patentType': rec.get('patentType', ''),
            'legalStatus': rec.get('legalStatus', '未知'),
            'source': rec.get('source', ''),
            'pubId': rec.get('pubId', ''),
        }
        # 如果类型仍为空，从标题推断
        if not item['patentType']:
            item['patentType'] = infer_patent_type(item['title'])
        # 计算过期日
        expiry, days = compute_expiry(item['applyDate'], item['patentType'], item['title'])
        item['expiryDate'] = expiry
        item['daysRemaining'] = days
        if item['applyDate'] and len(item['applyDate']) >= 4:
            item['applyYear'] = int(item['applyDate'][:4])
        else:
            item['applyYear'] = 0
        cleaned.append(item)
    return cleaned


def save_output(records, OUTPUT_DIR):
    OUTPUT_JSON = os.path.join(OUTPUT_DIR, '专利数据_清洗融合.json')
    OUTPUT_XLSX = os.path.join(OUTPUT_DIR, '专利数据_清洗融合.xlsx')
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    logger.info(f"JSON 已保存: {OUTPUT_JSON} ({len(records)} 条)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "专利数据"
    headers = ['申请号', '发明名称', '申请日', '公开(公告)日', '申请人',
               '公司', '发明人', 'IPC分类号', '地址', '专利类型',
               '法律状态', '剩余天数', '专利代理机构', '专利代理师', '摘要', '来源']
    hd_font = Font(bold=True, color='FFFFFF', size=11)
    hd_fill = PatternFill('solid', fgColor='4472C4')
    hd_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = hd_font; c.fill = hd_fill; c.alignment = hd_align; c.border = thin_border

    for ri, rec in enumerate(records, 2):
        vals = [
            f"CN{rec['applyId']}" if rec['applyId'] else '',
            rec['title'], rec['applyDate'], rec['pubDate'], rec['applicant'],
            rec['company'], rec['inventor'], rec['classification'], rec['address'],
            rec['patentType'], rec['legalStatus'], rec['daysRemaining'],
            rec['patentAgency'], rec['patentAgent'], rec['abstract'], rec['source']
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if isinstance(v, int) and v < 0:
                cell.font = Font(color='FF0000')

    col_widths = [18, 45, 12, 14, 25, 20, 30, 30, 30, 10, 10, 10, 25, 15, 60, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'
    wb.save(OUTPUT_XLSX)
    logger.info(f"Excel 已保存: {OUTPUT_XLSX} ({len(records)} 条)")


# ==================== 主流程 ====================
def main():
    parser = argparse.ArgumentParser(description='专利数据清洗融合')
    parser.add_argument('--data-dir', required=True, help='用户数据目录')
    args = parser.parse_args()

    data_dir = args.data_dir
    # 设置文件日志
    setup_file_logging(data_dir)

    logger.info("=" * 70)
    logger.info("专利数据清洗融合 v4.0 (Electron版)")
    logger.info(f"数据目录: {data_dir}")
    logger.info(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 70)

    INPUT_DIR = os.path.join(data_dir, "temp_data")
    OUTPUT_DIR = os.path.join(data_dir, 'cleaned_data')

    publish_dir = os.path.join(INPUT_DIR, '中国专利公布公告网')
    tianyan_dir = os.path.join(INPUT_DIR, '天眼查')
    search_dir = os.path.join(INPUT_DIR, '专利检索及分析网')

    logger.info("[1] 读取中国专利公布公告 CSV...")
    publish_data = read_publish_csv(publish_dir) if os.path.isdir(publish_dir) else []
    logger.info(f"    → {len(publish_data)} 条")

    logger.info("[2] 读取天眼查 XLSX...")
    tianyan_data = read_tianyan_xlsx(tianyan_dir) if os.path.isdir(tianyan_dir) else []
    logger.info(f"    → {len(tianyan_data)} 条")

    logger.info("[3] 读取专利检索分析网 XLSX...")
    search_data = read_search_xlsx(search_dir) if os.path.isdir(search_dir) else []
    logger.info(f"    → {len(search_data)} 条")

    logger.info("[4] 数据融合...")
    merged = merge_all(search_data, tianyan_data, publish_data)
    logger.info(f"    → 融合后 {len(merged)} 条")

    logger.info("[5] 最终清洗...")
    cleaned = final_clean(merged)

    logger.info("[6] 保存结果...")
    save_output(cleaned, OUTPUT_DIR)

    sources = {}
    types = {}
    for r in cleaned:
        s = r['source']
        sources[s] = sources.get(s, 0) + 1
        t = r['patentType']
        types[t] = types.get(t, 0) + 1
    logger.info("数据来源分布：")
    for k, v in sources.items():
        logger.info(f"  {k}: {v}")
    logger.info("专利类型分布：")
    for k, v in types.items():
        logger.info(f"  {k}: {v}")
    logger.info("完成！")


if __name__ == '__main__':
    main()